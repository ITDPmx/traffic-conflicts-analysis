
import openpyxl 
from openpyxl.styles import PatternFill
from openpyxl import Workbook

from pydantic import BaseModel
from typing import Dict, Optional
from metadata import Metadata, header_metadata

class EntityDetails(BaseModel):
    user_id: int
    sex_id: Optional[str]
    age: Optional[int]
    velocity: float
    colission_point_distance: float
    tc_value: float
    evasion_id: int
    evasion_possibility_id: int


class SmallEntityDetails(BaseModel):
    user_id: int
    sex_id: int
    age: int
    speed: float
 
 
class Event(BaseModel):
    video_name: str
    video_section: Optional[str]
    video_minute: int
    event_hour: str
    event_description: str
    entity1: EntityDetails
    entity2: EntityDetails
    entity3: SmallEntityDetails
    
class HeaderData(BaseModel):
    observer: Optional[str]
    city: str
    intersection: str
    weather: str
    date: str
    time_start: str
    time_end: str
    sunny: Optional[bool]
    dry: Optional[bool]
    cloudy: Optional[bool]
    humid: Optional[bool]
    surface: Optional[str]
    raining: Optional[bool]


def set_header(workbook: Workbook, header_data: HeaderData, header_metadata: Metadata):
    sheet = workbook.active
    header_data_dict = header_data.model_dump()
    
    for key, value in header_metadata.info.items():
        if key not in header_data_dict:
            continue
        
        target_row = value.row
        target_column = value.col
        
        try:
            if type(header_data_dict[key]) == bool:
                if header_data_dict[key]:
                    sheet.cell(row=target_row, column=target_column).fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
                else:
                    sheet.cell(row=target_row, column=target_column, value='X')
            else:
                sheet.cell(row=target_row, column=target_column, value=header_data_dict[key])
        except Exception as e:
            print(f"Error setting header data: {e}")
            print("Key: ", key, ", Value: ", header_data_dict[key], "Row: ", target_row, "Column: ", target_column)
            continue


def add_event(workbook: Workbook, event: Event, row: int, event_metadata: Metadata):
    pass


def add_events(workbook: Workbook, events: list[Event]):
    next_index = 2
    for event in events:
        next_index = add_event(workbook, event, next_index)



def main():
    test()

def get_sheet():
    path = "data/template.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active 
    return sheet_obj

def get_workbook():
    path = "data/template.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    return wb_obj

def test():
    header_data = HeaderData(
        observer="Alice",
        city="Wonderland",
        intersection="Rabbit Hole",
        weather="Sunny",
        sunny=True,
        dry=True,
        cloudy=False,
        humid=False,
        surface="Grass",
        raining=True,
        date="2021-10-10",
        time_start="10:00",
        time_end="10:30"
        )
    wb = get_workbook()
    set_header(wb, header_data, header_metadata)
    
    wb.save(filename="data/output_test.xlsx")
    
main()