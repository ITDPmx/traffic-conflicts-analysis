import openpyxl 
from openpyxl import Workbook
from openpyxl.styles import PatternFill

import json

from objects import HeaderData, Event, EntityDetails, SmallEntityDetails
from metadata import Metadata, header_metadata, event_details_offset, entity_column_offset

from metadata import validate_metadata, validate_offsets

def set_header(workbook: Workbook, header_data: HeaderData, header_metadata: Metadata):
    header_data_dict = header_data.model_dump()
    validate_metadata(header_data_dict, header_metadata)
    
    sheet = workbook.active
    
    for key, value in header_metadata.info.items():
        
        # Some values are optional, non-optional values are validated in pydantic schema
        if key not in header_data_dict:
            continue
        
        target_row = value.row
        target_column = value.col
        
        try:
            # Custom styling for boolean values
            if type(header_data_dict[key]) == bool:
                if header_data_dict[key]:
                    sheet.cell(row=target_row, column=target_column).fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
                else:
                    sheet.cell(row=target_row, column=target_column, value='X')
            else:
                # Place value as-is
                sheet.cell(row=target_row, column=target_column, value=header_data_dict[key])
        except Exception as e:
            print(f"Error setting header data: {e}")
            print("Key:", key + ", Value:", header_data_dict[key] + ", Row:", target_row, + ", Column:", target_column)


def add_event(workbook: Workbook, event: Event, row: int):
    event_dict = event.model_dump()
    validate_offsets(event_dict, event_details_offset)
    
    sheet = workbook.active
    
    for offset in event_details_offset.keys():
        if offset == "entity1" or offset == "entity2":
            add_entity(workbook, EntityDetails(**event_dict[offset]), row, event_details_offset[offset])
        elif offset == "entity3":
            add_small_entity(workbook, event_dict[offset], row, event_details_offset[offset])
        else:
            cell = sheet.cell(row=row, column=event_details_offset[offset])
            cell.value = event_dict[offset]    
         
    return row + 1


def add_entity(workbook: Workbook, entity: EntityDetails, row: int, col: int):
    sheet = workbook.active
    entity_dict = entity.model_dump()
    
    for offset in entity_column_offset.keys():
        cell = sheet.cell(row=row, column=col + entity_column_offset[offset])
        cell.value = entity_dict[offset]


def add_small_entity(workbook: Workbook, entity: SmallEntityDetails, row: int, col: int):
    pass


def add_events(workbook: Workbook, events: list[Event]):
    next_index = 10
    for event in events:
        next_index = add_event(workbook, event, next_index)


def get_template():
    path = "data/template.xlsx"
    # path = "data/output_test.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    return wb_obj


def get_starting_row(workbook: Workbook):
    sheet = workbook.active
    
    for i in range(1, sheet.max_row):
                
        if sheet.cell(row=i, column=1).value is None:
            return i
        elif sheet.cell(row=i, column=1).merged_cells:
            for merged_range in sheet.merged_cells.ranges:
                if i >= merged_range.min_row and i <= merged_range.max_row:
                    return merged_range.max_row

def test():
    
    with open('data/mock.json') as json_data:
        test_data = json.load(json_data)
    
    header_data = test_data["header_data"]
    
    header_data = HeaderData(**header_data)
    
    wb = get_template()
    # print(get_starting_row(wb))
    
    set_header(wb, header_data, header_metadata)
    
    t = test_data["events"][0]
    
    events = [Event(**event) for event in test_data["events"]]
    add_events(wb, events)
    
    wb.save(filename="data/output_test.xlsx")
    

def main():
    test()

if __name__ == "__main__":
    main()